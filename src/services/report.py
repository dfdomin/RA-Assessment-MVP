from __future__ import annotations

from collections import defaultdict
from html import escape
from io import BytesIO
from zipfile import ZIP_DEFLATED, ZipFile
from typing import Any

import openpyxl
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from src.models.action_plan import ActionPlan
from src.models.assessment import Assessment
from src.models.leader_analysis import LeaderAnalysis
from src.models.leader_report import LeaderReportDraft
from src.models.module import Module, ModuleAssignment
from src.models.module_analysis import ModuleAnalysis
from src.models.period import Period
from src.models.rubric import PILevel, PerfIndicator
from src.models.student import ModuleStudent, Student
from src.models.student_outcome import StudentOutcome
from src.models.user import User
from src.services.sanitize import safe_cell_value

LEVEL_LABELS = {1: "Poor", 2: "Inadequate", 3: "Adequate", 4: "Exemplary"}
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


async def active_pis_for_period(period: Period, db: AsyncSession) -> list[PerfIndicator]:
    if not period.rubric_id:
        return []

    result = await db.execute(
        select(PerfIndicator)
        .where(PerfIndicator.rubric_id == period.rubric_id, PerfIndicator.is_active.is_(True))
        .order_by(PerfIndicator.position)
    )
    return list(result.scalars().all())


async def build_report_data(period: Period, db: AsyncSession) -> dict[str, Any]:
    so = await db.get(StudentOutcome, period.student_outcome_id)
    active_pis = await active_pis_for_period(period, db)
    pi_by_id = {pi.id: pi for pi in active_pis}

    module_rows = await _module_rows(period.id, db)
    modules_summary = [
        {
            "module_id": row["module"].id,
            "course_code": row["module"].course_code,
            "course_name": row["module"].course_name,
            "group_name": row["module"].group_name,
            "teacher_names": row["teacher_names"],
            "active_students": row["active_students"],
        }
        for row in module_rows
        if row["active_students"] > 0
    ]

    level_descriptors = await _level_descriptors(pi_by_id, db)
    distribution = await _distribution_by_pi(period.id, pi_by_id, db)
    leader_analysis = await _leader_analysis_by_pi(period.id, pi_by_id, db)
    teacher_analysis = await _teacher_analysis_by_pi(period.id, pi_by_id, db)
    action_plans = await _action_plans(period.id, pi_by_id, db)

    return {
        "period": {
            "id": period.id,
            "name": period.name,
            "status": period.status,
            "start_date": period.start_date.isoformat(),
            "end_date": period.end_date.isoformat(),
        },
        "student_outcome": {
            "id": so.id if so else None,
            "code": so.code if so else None,
            "description": so.description if so else None,
        },
        "modules_summary": modules_summary,
        "distribution_by_pi": {
            pi.code: {
                "perf_indicator_id": pi.id,
                "description": pi.description,
                "levels": level_descriptors.get(pi.id, []),
                "by_module": distribution.get(pi.id, {}).get("by_module", []),
                "consolidated": distribution.get(pi.id, {}).get("consolidated", _empty_distribution()),
            }
            for pi in active_pis
            if distribution.get(pi.id, {}).get("total", 0) > 0
        },
        "leader_analysis": leader_analysis,
        "teacher_analysis": teacher_analysis,
        "action_plans": action_plans,
    }


async def missing_export_prerequisites(
    period: Period, db: AsyncSession
) -> dict[str, list[str]]:
    active_pis = await active_pis_for_period(period, db)
    pi_by_id = {pi.id: pi for pi in active_pis}

    la_result = await db.execute(select(LeaderAnalysis).where(LeaderAnalysis.period_id == period.id))
    complete_la_ids = {
        item.perf_indicator_id for item in la_result.scalars().all() if item.analysis_text.strip()
    }

    ap_result = await db.execute(select(ActionPlan).where(ActionPlan.period_id == period.id))
    complete_ap_ids = {
        item.perf_indicator_id
        for item in ap_result.scalars().all()
        if item.description.strip() and item.responsible.strip() and item.estimated_date.strip()
    }

    return {
        "missing_leader_analysis": [
            pi.code for pi_id, pi in pi_by_id.items() if pi_id not in complete_la_ids
        ],
        "missing_action_plans": [
            pi.code for pi_id, pi in pi_by_id.items() if pi_id not in complete_ap_ids
        ],
    }


def render_pdf(report: dict[str, Any]) -> bytes:
    html = _render_html(report)
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        return _minimal_pdf(html)

    try:
        return HTML(string=html).write_pdf()
    except OSError:
        return _minimal_pdf(html)


def render_xlsx(report: dict[str, Any]) -> bytes:
    workbook = openpyxl.Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    _append_safe(summary, ["Periodo", report["period"]["name"]])
    _append_safe(summary, ["RA", report["student_outcome"]["code"], report["student_outcome"]["description"]])
    _append_safe(summary, [])
    _append_safe(summary, ["Curso", "Grupo", "Docentes", "Estudiantes activos"])
    for module in report["modules_summary"]:
        _append_safe(
            summary,
            [
                module["course_code"],
                module["group_name"],
                ", ".join(module["teacher_names"]),
                module["active_students"],
            ],
        )

    distribution = workbook.create_sheet("Distribucion")
    _append_safe(
        distribution,
        ["PI", "Descripcion PI", "Nivel", "Descriptor", "Modulo", "Porcentaje", "Conteo"],
    )
    for pi_code, pi_data in report["distribution_by_pi"].items():
        descriptors = {
            level["label"]: level["descriptor"]
            for level in pi_data.get("levels", [])
        }
        for module_row in pi_data["by_module"]:
            counts = module_row["counts"]
            percentages = module_row["percentages"]
            for level in LEVEL_LABELS.values():
                _append_safe(
                    distribution,
                    [
                        pi_code,
                        pi_data["description"],
                        level,
                        descriptors.get(level, ""),
                        module_row["course_code"],
                        percentages[level],
                        counts[level],
                    ],
                )
        consolidated = pi_data["consolidated"]
        for level in LEVEL_LABELS.values():
            _append_safe(
                distribution,
                [
                    pi_code,
                    pi_data["description"],
                    level,
                    descriptors.get(level, ""),
                    "TOTAL CONSOLIDADO",
                    consolidated["percentages"][level],
                    consolidated["counts"][level],
                ],
            )

    analysis = workbook.create_sheet("Analisis")
    _append_safe(analysis, ["PI", "Modulo", "Analisis docente", "Sintesis lider"])
    for pi_code, leader_text in report["leader_analysis"].items():
        teacher_items = report["teacher_analysis"].get(pi_code, [])
        if not teacher_items:
            _append_safe(analysis, [pi_code, "", "", leader_text])
        for item in teacher_items:
            _append_safe(
                analysis,
                [pi_code, item["course_code"], item["analysis_text"], leader_text],
            )

    plans = workbook.create_sheet("Plan de Accion")
    _append_safe(plans, ["PI", "Standard", "Tipo", "Descripcion", "Responsable", "Fecha estimada"])
    for plan in report["action_plans"]:
        _append_safe(
            plans,
            [
                plan["pi_code"],
                plan["standard"],
                plan["action_type"],
                plan["description"],
                plan["responsible"],
                plan["estimated_date"],
            ],
        )

    students = workbook.create_sheet("Estudiantes")
    _append_safe(students, ["Curso", "Nombre estudiante"])
    for row in report.get("student_names", []):
        _append_safe(students, [row["course_code"], row["student_name"]])

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


async def build_leader_report_data(period: Period, db: AsyncSession) -> dict[str, Any]:
    report = await build_report_data(period, db)
    active_pis = await active_pis_for_period(period, db)

    draft_result = await db.execute(
        select(LeaderReportDraft).where(LeaderReportDraft.period_id == period.id)
    )
    drafts = {item.perf_indicator_id: item.conclusion_text for item in draft_result.scalars().all()}
    action_plans = {
        item["perf_indicator_id"]: item for item in report["action_plans"]
    }

    items: list[dict[str, Any]] = []
    for pi in active_pis:
        distribution = report["distribution_by_pi"].get(pi.code, _empty_distribution())
        items.append(
            {
                "perf_indicator_id": pi.id,
                "pi_code": pi.code,
                "pi_description": pi.description,
                "distribution": distribution.get("consolidated", _empty_distribution()),
                "teacher_analysis": report["teacher_analysis"].get(pi.code, []),
                "leader_analysis": report["leader_analysis"].get(pi.code, ""),
                "action_plan": action_plans.get(pi.id),
                "conclusion_text": drafts.get(pi.id, ""),
            }
        )

    return {
        "period": report["period"],
        "student_outcome": report["student_outcome"],
        "items": items,
    }


def render_leader_report_pdf(leader_report: dict[str, Any]) -> bytes:
    html = _render_leader_report_html(leader_report)
    try:
        from weasyprint import HTML
    except (ImportError, OSError):
        return _minimal_pdf(html)

    try:
        return HTML(string=html).write_pdf()
    except OSError:
        return _minimal_pdf(html)


def render_leader_report_docx(leader_report: dict[str, Any]) -> bytes:
    body = "".join(
        _docx_paragraph(f"{item['pi_code']} - {safe_cell_value(item['pi_description'])}", "Heading2")
        + _docx_paragraph(f"Analisis lider: {safe_cell_value(item['leader_analysis'])}")
        + _docx_paragraph(f"Conclusion: {safe_cell_value(item['conclusion_text'])}")
        for item in leader_report["items"]
    )
    document_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">'
        f"<w:body>{body}<w:sectPr><w:pgSz w:w=\"12240\" w:h=\"15840\"/>"
        '<w:pgMar w:top="1440" w:right="1440" w:bottom="1440" w:left="1440"/>'
        "</w:sectPr></w:body></w:document>"
    )

    output = BytesIO()
    with ZipFile(output, "w", ZIP_DEFLATED) as docx:
        docx.writestr(
            "[Content_Types].xml",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">'
            '<Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/>'
            '<Default Extension="xml" ContentType="application/xml"/>'
            '<Override PartName="/word/document.xml" '
            'ContentType="application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml"/>'
            "</Types>",
        )
        docx.writestr(
            "_rels/.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" '
            'Target="word/document.xml"/></Relationships>',
        )
        docx.writestr("word/document.xml", document_xml)
    return output.getvalue()


async def _module_rows(period_id: int, db: AsyncSession) -> list[dict[str, Any]]:
    result = await db.execute(
        select(Module)
        .where(Module.period_id == period_id)
        .order_by(Module.course_code, Module.group_name)
    )
    modules = list(result.scalars().all())
    rows: list[dict[str, Any]] = []

    for module in modules:
        count_result = await db.execute(
            select(ModuleStudent).where(
                ModuleStudent.module_id == module.id,
                ModuleStudent.status == "active",
            )
        )
        active_students = len(count_result.scalars().all())
        teacher_result = await db.execute(
            select(User.full_name)
            .join(ModuleAssignment, ModuleAssignment.user_id == User.id)
            .where(ModuleAssignment.module_id == module.id)
            .order_by(User.full_name)
        )
        rows.append(
            {
                "module": module,
                "active_students": active_students,
                "teacher_names": list(teacher_result.scalars().all()),
            }
        )
    return rows


async def _level_descriptors(
    pi_by_id: dict[int, PerfIndicator], db: AsyncSession
) -> dict[int, list[dict[str, Any]]]:
    if not pi_by_id:
        return {}
    result = await db.execute(
        select(PILevel)
        .where(PILevel.perf_indicator_id.in_(pi_by_id))
        .order_by(PILevel.perf_indicator_id, PILevel.level_value)
    )
    grouped: dict[int, list[dict[str, Any]]] = defaultdict(list)
    for level in result.scalars().all():
        grouped[level.perf_indicator_id].append(
            {
                "level": level.level_value,
                "label": level.label,
                "descriptor": level.descriptor,
            }
        )
    return grouped


async def _distribution_by_pi(
    period_id: int, pi_by_id: dict[int, PerfIndicator], db: AsyncSession
) -> dict[int, dict[str, Any]]:
    result = await db.execute(
        select(
            PerfIndicator.id,
            Module.id,
            Module.course_code,
            Assessment.level,
        )
        .join(Assessment, Assessment.perf_indicator_id == PerfIndicator.id)
        .join(ModuleStudent, Assessment.module_student_id == ModuleStudent.id)
        .join(Module, ModuleStudent.module_id == Module.id)
        .where(
            Module.period_id == period_id,
            ModuleStudent.status == "active",
            PerfIndicator.id.in_(pi_by_id) if pi_by_id else False,
        )
    )
    grouped: dict[int, dict[int, dict[str, Any]]] = defaultdict(
        lambda: defaultdict(lambda: {"counts": _empty_counts(), "course_code": ""})
    )
    consolidated: dict[int, dict[str, int]] = defaultdict(_empty_counts)

    for pi_id, module_id, course_code, level in result.all():
        label = LEVEL_LABELS.get(level, "Poor")
        grouped[pi_id][module_id]["course_code"] = course_code
        grouped[pi_id][module_id]["counts"][label] += 1
        consolidated[pi_id][label] += 1

    data: dict[int, dict[str, Any]] = {}
    for pi_id, module_data in grouped.items():
        total = sum(consolidated[pi_id].values())
        data[pi_id] = {
            "total": total,
            "by_module": [
                {
                    "module_id": module_id,
                    "course_code": values["course_code"],
                    "counts": values["counts"],
                    "percentages": _percentages(values["counts"]),
                }
                for module_id, values in module_data.items()
            ],
            "consolidated": {
                "counts": consolidated[pi_id],
                "percentages": _percentages(consolidated[pi_id]),
            },
        }
    return data


async def _leader_analysis_by_pi(
    period_id: int, pi_by_id: dict[int, PerfIndicator], db: AsyncSession
) -> dict[str, str]:
    result = await db.execute(select(LeaderAnalysis).where(LeaderAnalysis.period_id == period_id))
    values: dict[str, str] = {}
    for item in result.scalars().all():
        pi = pi_by_id.get(item.perf_indicator_id)
        if pi:
            values[pi.code] = item.analysis_text
    return values


async def _teacher_analysis_by_pi(
    period_id: int, pi_by_id: dict[int, PerfIndicator], db: AsyncSession
) -> dict[str, list[dict[str, str]]]:
    result = await db.execute(
        select(PerfIndicator.id, Module.course_code, ModuleAnalysis.analysis_text)
        .join(ModuleAnalysis, ModuleAnalysis.perf_indicator_id == PerfIndicator.id)
        .join(Module, ModuleAnalysis.module_id == Module.id)
        .where(Module.period_id == period_id, PerfIndicator.id.in_(pi_by_id) if pi_by_id else False)
        .order_by(PerfIndicator.position, Module.course_code)
    )
    values: dict[str, list[dict[str, str]]] = defaultdict(list)
    for pi_id, course_code, analysis_text in result.all():
        pi = pi_by_id.get(pi_id)
        if pi:
            values[pi.code].append({"course_code": course_code, "analysis_text": analysis_text})
    return values


async def _action_plans(
    period_id: int, pi_by_id: dict[int, PerfIndicator], db: AsyncSession
) -> list[dict[str, Any]]:
    result = await db.execute(select(ActionPlan).where(ActionPlan.period_id == period_id))
    plans = []
    for item in result.scalars().all():
        pi = pi_by_id.get(item.perf_indicator_id)
        if pi:
            plans.append(
                {
                    "perf_indicator_id": item.perf_indicator_id,
                    "pi_code": pi.code,
                    "standard": _standard_from_action_type(item.action_type),
                    "action_type": item.action_type,
                    "description": item.description,
                    "responsible": item.responsible,
                    "estimated_date": item.estimated_date,
                    "implemented": item.implemented,
                }
            )
    return sorted(plans, key=lambda item: item["pi_code"])


async def student_names_for_period(period_id: int, db: AsyncSession) -> list[dict[str, str]]:
    result = await db.execute(
        select(Module.course_code, Student.full_name)
        .join(ModuleStudent, ModuleStudent.module_id == Module.id)
        .join(Student, ModuleStudent.student_id == Student.id)
        .where(Module.period_id == period_id, ModuleStudent.status == "active")
        .order_by(Module.course_code, Student.full_name)
    )
    return [{"course_code": course_code, "student_name": full_name} for course_code, full_name in result.all()]


def _empty_counts() -> dict[str, int]:
    return {"Poor": 0, "Inadequate": 0, "Adequate": 0, "Exemplary": 0}


def _empty_distribution() -> dict[str, Any]:
    return {"counts": _empty_counts(), "percentages": _percentages(_empty_counts())}


def _percentages(counts: dict[str, int]) -> dict[str, float]:
    total = sum(counts.values())
    if total == 0:
        return {label: 0.0 for label in LEVEL_LABELS.values()}
    return {label: round((counts[label] / total) * 100, 2) for label in LEVEL_LABELS.values()}


def _standard_from_action_type(action_type: str) -> str:
    if action_type == "corrective":
        return "Low"
    if action_type == "improvement":
        return "High"
    return "Medium"


def _append_safe(sheet: openpyxl.worksheet.worksheet.Worksheet, row: list[Any]) -> None:
    sheet.append([safe_cell_value(value) for value in row])


def _render_html(report: dict[str, Any]) -> str:
    modules = "".join(
        f"<tr><td>{escape(module['course_code'])}</td><td>{escape(module['group_name'])}</td>"
        f"<td>{escape(', '.join(module['teacher_names']))}</td><td>{module['active_students']}</td></tr>"
        for module in report["modules_summary"]
    )
    plans = "".join(
        f"<tr><td>{escape(plan['pi_code'])}</td><td>{escape(plan['standard'])}</td>"
        f"<td>{escape(plan['action_type'])}</td><td>{escape(plan['description'])}</td>"
        f"<td>{escape(plan['responsible'])}</td><td>{escape(plan['estimated_date'])}</td></tr>"
        for plan in report["action_plans"]
    )
    return f"""
    <html>
      <head><meta charset="utf-8"><title>Reporte ABET</title></head>
      <body>
        <h1>Reporte ABET {escape(report['period']['name'])}</h1>
        <h2>Seccion 1 - Encabezado</h2>
        <p>{escape(str(report['student_outcome']['code']))}: {escape(str(report['student_outcome']['description']))}</p>
        <table>{modules}</table>
        <h2>Seccion 2 - Distribucion por PI</h2>
        <pre>{escape(str(report['distribution_by_pi']))}</pre>
        <h2>Seccion 3 - Analisis del lider</h2>
        <pre>{escape(str(report['leader_analysis']))}</pre>
        <h2>Seccion 4 - Plan de Accion</h2>
        <table>{plans}</table>
      </body>
    </html>
    """


def _render_leader_report_html(leader_report: dict[str, Any]) -> str:
    items = "".join(
        f"<section><h2>{escape(item['pi_code'])}</h2>"
        f"<p>{escape(item['pi_description'])}</p>"
        f"<h3>Analisis del lider</h3><p>{escape(item['leader_analysis'])}</p>"
        f"<h3>Conclusion</h3><p>{escape(item['conclusion_text'])}</p></section>"
        for item in leader_report["items"]
    )
    return f"""
    <html>
      <head><meta charset="utf-8"><title>Informe del Lider</title></head>
      <body>
        <h1>Informe del Lider {escape(leader_report['period']['name'])}</h1>
        <p>{escape(str(leader_report['student_outcome']['code']))}: {escape(str(leader_report['student_outcome']['description']))}</p>
        {items}
      </body>
    </html>
    """


def _docx_paragraph(text: Any, style: str | None = None) -> str:
    style_xml = f'<w:pPr><w:pStyle w:val="{escape(style)}"/></w:pPr>' if style else ""
    text_xml = escape(str(text)).replace("&#x27;", "'")
    return f"<w:p>{style_xml}<w:r><w:t>{text_xml}</w:t></w:r></w:p>"


def _minimal_pdf(text: str) -> bytes:
    safe_text = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 12 Tf 72 720 Td ({safe_text[:1200]}) Tj ET".encode("latin-1", errors="ignore")
    objects = [
        b"1 0 obj << /Type /Catalog /Pages 2 0 R >> endobj\n",
        b"2 0 obj << /Type /Pages /Kids [3 0 R] /Count 1 >> endobj\n",
        b"3 0 obj << /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >> endobj\n",
        b"4 0 obj << /Type /Font /Subtype /Type1 /BaseFont /Helvetica >> endobj\n",
        b"5 0 obj << /Length " + str(len(stream)).encode() + b" >> stream\n" + stream + b"\nendstream endobj\n",
    ]
    pdf = bytearray(b"%PDF-1.4\n")
    offsets = [0]
    for obj in objects:
        offsets.append(len(pdf))
        pdf.extend(obj)
    xref_start = len(pdf)
    pdf.extend(f"xref\n0 {len(objects) + 1}\n0000000000 65535 f \n".encode())
    for offset in offsets[1:]:
        pdf.extend(f"{offset:010d} 00000 n \n".encode())
    pdf.extend(
        f"trailer << /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_start}\n%%EOF\n".encode()
    )
    return bytes(pdf)

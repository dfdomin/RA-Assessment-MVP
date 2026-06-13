#!/usr/bin/env python3
"""
Backfill modules.program_id from MODULOS {cuatrimestre} POR RESULTADOS DE APRENDIZAJE.xlsx.

Matches each physical module by (course_code, group_name) → PROGRAMA column in the Excel.
Requires service role key for updates.

Usage:
  python3 scripts/backfill_module_program_ids_from_mapping.py --dry-run
  python3 scripts/backfill_module_program_ids_from_mapping.py
  python3 scripts/backfill_module_program_ids_from_mapping.py --excel path/to/file.xlsx --cycle 2025-2
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import unicodedata
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "MODULOS 2025-2 POR RESULTADOS DE APRENDIZAJE.xlsx"
DEFAULT_CYCLE = "2025-2"
DEFAULT_SUPABASE_URL = "https://whjjervbojyktkhvvmte.supabase.co"
MAPPING_SHEETS = ("RA CE TGLI ANI", "RA TGA INE")

# Keep in sync with scripts/seed_consolidators_from_mapping.py
PROGRAM_ALIASES: dict[str, list[str]] = {
    "TGA": ["TG Administrativa", "Tecnología en Gestión Administrativa", "TGA"],
    "ING-NEGOCIOS": [
        "Inteligencia de Negocios",
        "Profesional en Inteligencia de Negocios",
    ],
    "CE": ["Comercio Exterior"],
    "TGLI": ["TG Logística Internacional", "Tecnología en Gestión Logística Internacional"],
    "ANI": ["Adm. Neg. Internacionales", "Adm. Negocios Internacionales"],
}


@dataclass(frozen=True)
class ModuleMappingRow:
    course_code: str
    group_name: str
    program_label: str
    sheet: str


def load_env_file(path: Path) -> None:
    if not path.is_file():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def normalize_name(value: str) -> str:
    text = unicodedata.normalize("NFKD", value or "")
    text = "".join(c for c in text if not unicodedata.combining(c))
    return " ".join(text.strip().upper().split())


def program_code_for_label(programs: list[dict], label: str) -> str | None:
    norm_label = normalize_name(label)
    for code, aliases in PROGRAM_ALIASES.items():
        for alias in aliases:
            if normalize_name(alias) == norm_label:
                for prog in programs:
                    if prog.get("code") == code:
                        return code
    for prog in programs:
        if normalize_name(prog.get("name", "")) == norm_label:
            return prog["code"]
    return None


def parse_mapping_excel(path: Path) -> list[ModuleMappingRow]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("pip install openpyxl") from exc

    if not path.is_file():
        raise SystemExit(f"Excel not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[ModuleMappingRow] = []
    seen: set[tuple[str, str]] = set()

    for sheet_name in MAPPING_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Missing sheet {sheet_name!r} in {path.name}")
        ws = wb[sheet_name]
        for raw in ws.iter_rows(min_row=4, values_only=True):
            if not raw or raw[2] is None or raw[3] is None:
                continue
            group_name = str(raw[2]).strip()
            course_code = str(raw[3]).strip()
            program_label = str(raw[6]).strip() if raw[6] else ""
            if not group_name or not course_code or not program_label:
                continue
            key = (course_code, group_name)
            if key in seen:
                raise SystemExit(
                    f"Duplicate (CODIGO, GRUPO) in Excel: {course_code} · {group_name}"
                )
            seen.add(key)
            rows.append(
                ModuleMappingRow(
                    course_code=course_code,
                    group_name=group_name,
                    program_label=program_label,
                    sheet=sheet_name,
                )
            )

    wb.close()
    return rows


def build_program_lookup(
    mapping_rows: list[ModuleMappingRow], programs: list[dict]
) -> tuple[dict[tuple[str, str], int], list[str]]:
    programs_by_code = {p["code"]: p for p in programs}
    lookup: dict[tuple[str, str], int] = {}
    unresolved: list[str] = []

    for row in mapping_rows:
        code = program_code_for_label(programs, row.program_label)
        if not code or code not in programs_by_code:
            unresolved.append(f"{row.course_code}|{row.group_name}|program:{row.program_label}")
            continue
        lookup[(row.course_code, row.group_name)] = programs_by_code[code]["id"]

    return lookup, unresolved


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Report only; do not update Supabase")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to mapping Excel")
    parser.add_argument("--cycle", default=DEFAULT_CYCLE, help="measurement_cycles.code (default: 2025-2)")
    parser.add_argument(
        "--export-json",
        type=Path,
        default=None,
        help="Optional path to write parsed module→program map as JSON",
    )
    args = parser.parse_args()

    try:
        from supabase import create_client
    except ImportError:
        print("pip install supabase", file=sys.stderr)
        return 1

    load_env_file(ROOT / ".env")
    load_env_file(ROOT / ".env.local")

    url = os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL)
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SECRET_KEY")
    if not key:
        print(
            "Falta SUPABASE_SERVICE_ROLE_KEY en .env\n"
            "Supabase Dashboard → Project Settings → API → service_role",
            file=sys.stderr,
        )
        return 1

    mapping_rows = parse_mapping_excel(args.excel)
    print(f"Excel rows parsed: {len(mapping_rows)} from {args.excel.name}")

    sb = create_client(url, key)
    cycle = (
        sb.table("measurement_cycles").select("id, code").eq("code", args.cycle).single().execute()
    ).data
    programs = sb.table("programs").select("id, code, name").execute().data or []
    program_lookup, unresolved_labels = build_program_lookup(mapping_rows, programs)
    if unresolved_labels:
        print(f"Unresolved program labels ({len(unresolved_labels)}):")
        for item in unresolved_labels[:20]:
            print(f"  - {item}")
        return 1

    modules = (
        sb.table("modules")
        .select("id, course_code, group_name, program_id, cycle_id")
        .eq("cycle_id", cycle["id"])
        .execute()
    ).data or []

    to_update: list[dict] = []
    already_ok = 0
    unmatched_db: list[str] = []
    conflicts: list[str] = []

    db_by_key = {(m["course_code"], m["group_name"]): m for m in modules}
    excel_keys = set(program_lookup.keys())

    for key, program_id in sorted(program_lookup.items()):
        mod = db_by_key.get(key)
        if not mod:
            unmatched_db.append(f"{key[0]} · {key[1]}")
            continue
        current = mod.get("program_id")
        if current == program_id:
            already_ok += 1
            continue
        if current is not None and current != program_id:
            conflicts.append(
                f"{key[0]} · {key[1]}: DB program_id={current}, Excel→{program_id}"
            )
            continue
        to_update.append({"id": mod["id"], "program_id": program_id, "key": key})

    unmatched_excel = sorted(excel_keys - set(db_by_key.keys()))

    print(f"Cycle: {cycle['code']} (id={cycle['id']})")
    print(f"Modules in DB: {len(modules)}")
    print(f"Already correct program_id: {already_ok}")
    print(f"To update (NULL → program): {len(to_update)}")
    print(f"Excel rows without DB module: {len(unmatched_excel)}")
    print(f"DB modules without Excel row: {len(unmatched_db)}")

    if unmatched_excel[:10]:
        print("Sample Excel-only keys:")
        for code, group in unmatched_excel[:10]:
            print(f"  - {code} · {group}")

    if unmatched_db[:10]:
        print("Sample DB-only keys:")
        for item in unmatched_db[:10]:
            print(f"  - {item}")

    if conflicts:
        print(f"Conflicts ({len(conflicts)}) — manual review required:")
        for item in conflicts[:20]:
            print(f"  - {item}")
        return 1

    if args.export_json:
        payload = {
            "cycle_code": cycle["code"],
            "source_excel": str(args.excel),
            "modules": [
                {
                    "course_code": row.course_code,
                    "group_name": row.group_name,
                    "program_label": row.program_label,
                    "program_id": program_lookup[(row.course_code, row.group_name)],
                    "sheet": row.sheet,
                }
                for row in mapping_rows
            ],
        }
        args.export_json.parent.mkdir(parents=True, exist_ok=True)
        args.export_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Exported JSON: {args.export_json}")

    if args.dry_run:
        print("\n[dry-run] Sample updates:")
        for row in to_update[:15]:
            print(f"  module {row['id']}: {row['key'][0]} · {row['key'][1]} → program_id={row['program_id']}")
        return 0

    updated = 0
    for row in to_update:
        sb.table("modules").update({"program_id": row["program_id"]}).eq("id", row["id"]).execute()
        updated += 1

    print(f"Updated modules.program_id: {updated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

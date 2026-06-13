#!/usr/bin/env python3
"""
Sync module_ra_evaluations from MODULOS {cuatrimestre} POR RESULTADOS DE APRENDIZAJE.xlsx.

Creates missing module × RA (period) rows from Excel X marks in RA1–RA6 columns.
Requires service role key for inserts.

Usage:
  python3 scripts/sync_module_ra_evaluations_from_mapping.py --dry-run
  python3 scripts/sync_module_ra_evaluations_from_mapping.py
  python3 scripts/sync_module_ra_evaluations_from_mapping.py --excel path/to/file.xlsx --cycle 2025-2
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_EXCEL = ROOT / "MODULOS 2025-2 POR RESULTADOS DE APRENDIZAJE.xlsx"
DEFAULT_CYCLE = "2025-2"
DEFAULT_SUPABASE_URL = "https://whjjervbojyktkhvvmte.supabase.co"
MAPPING_SHEETS = ("RA CE TGLI ANI", "RA TGA INE")
RA_COLUMNS: tuple[tuple[int, str], ...] = (
    (7, "RA1"),
    (8, "RA2"),
    (9, "RA3"),
    (10, "RA4"),
    (11, "RA5"),
    (12, "RA6"),
)


@dataclass(frozen=True)
class ModuleRaAssignment:
    course_code: str
    group_name: str
    ra_code: str
    program_label: str
    sheet: str


def load_env_file(path: Path) -> None:
    if not path.is_file():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, _, value = line.partition("=")
        key = key.strip()
        value = value.strip().strip('"').strip("'")
        if key and key not in os.environ:
            os.environ[key] = value


def parse_ra_assignments(path: Path) -> list[ModuleRaAssignment]:
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit("pip install openpyxl") from exc

    if not path.is_file():
        raise SystemExit(f"Excel not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[ModuleRaAssignment] = []
    seen: set[tuple[str, str, str]] = set()

    for sheet_name in MAPPING_SHEETS:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(f"Missing sheet {sheet_name!r} in {path.name}")
        ws = wb[sheet_name]
        for raw in ws.iter_rows(min_row=4, values_only=True):
            if not raw or raw[2] is None or raw[3] is None:
                continue
            group_name = str(raw[2]).strip()
            course_code = str(raw[3]).strip()
            program_label = str(raw[6]).strip() if raw[6] else ""
            if not group_name or not course_code:
                continue
            for col_idx, ra_code in RA_COLUMNS:
                val = raw[col_idx] if col_idx < len(raw) else None
                if not val or str(val).strip().upper() != "X":
                    continue
                key = (course_code, group_name, ra_code)
                if key in seen:
                    raise SystemExit(
                        f"Duplicate assignment in Excel: {course_code} · {group_name} · {ra_code}"
                    )
                seen.add(key)
                rows.append(
                    ModuleRaAssignment(
                        course_code=course_code,
                        group_name=group_name,
                        ra_code=ra_code,
                        program_label=program_label,
                        sheet=sheet_name,
                    )
                )

    wb.close()
    return rows


def build_period_by_ra(periods: list[dict]) -> dict[str, int]:
    lookup: dict[str, int] = {}
    for period in periods:
        so = period.get("student_outcome") or period.get("student_outcomes")
        code = (so or {}).get("code")
        if code:
            lookup[code] = period["id"]
    return lookup


def build_module_lookup(modules: list[dict]) -> dict[tuple[str, str], dict]:
    lookup: dict[tuple[str, str], dict] = {}
    for mod in modules:
        key = (mod["course_code"], mod["group_name"])
        if key in lookup:
            raise SystemExit(
                f"Duplicate module in DB for {key[0]} · {key[1]}: "
                f"ids {lookup[key]['id']} and {mod['id']}"
            )
        lookup[key] = mod
    return lookup


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true", help="Report only; do not insert")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL, help="Path to mapping Excel")
    parser.add_argument("--cycle", default=DEFAULT_CYCLE, help="measurement_cycles.code (default: 2025-2)")
    parser.add_argument(
        "--export-json",
        type=Path,
        default=None,
        help="Optional path to write planned inserts as JSON",
    )
    args = parser.parse_args()

    try:
        from supabase import create_client
    except ImportError:
        print("pip install supabase", file=sys.stderr)
        return 1

    load_env_file(ROOT / ".env")
    load_env_file(ROOT / ".env.local")

    url = os.environ.get("SUPABASE_URL", DEFAULT_SUPABASE_URL)
    key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get("SUPABASE_SECRET_KEY")
    if not key:
        print(
            "Falta SUPABASE_SERVICE_ROLE_KEY en .env\n"
            "Supabase Dashboard → Project Settings → API → service_role",
            file=sys.stderr,
        )
        return 1

    assignments = parse_ra_assignments(args.excel)
    print(f"Excel module×RA assignments: {len(assignments)}")
    print(f"By RA: {dict(Counter(a.ra_code for a in assignments))}")

    sb = create_client(url, key)
    cycle = (
        sb.table("measurement_cycles").select("id, code").eq("code", args.cycle).single().execute()
    ).data
    periods = (
        sb.table("periods")
        .select("id, name, student_outcome_id, student_outcome:student_outcomes(code)")
        .eq("cycle_id", cycle["id"])
        .execute()
    ).data or []
    modules = (
        sb.table("modules")
        .select("id, course_code, group_name, program_id")
        .eq("cycle_id", cycle["id"])
        .execute()
    ).data or []
    existing = (
        sb.table("module_ra_evaluations")
        .select("id, module_id, period_id, status")
        .execute()
    ).data or []

    period_by_ra = build_period_by_ra(periods)
    module_by_key = build_module_lookup(modules)
    existing_pairs = {(row["module_id"], row["period_id"]) for row in existing}

    missing_module: list[str] = []
    missing_period: list[str] = []
    to_insert: list[dict] = []

    for assign in assignments:
        mod = module_by_key.get((assign.course_code, assign.group_name))
        if not mod:
            missing_module.append(f"{assign.course_code} · {assign.group_name} · {assign.ra_code}")
            continue
        period_id = period_by_ra.get(assign.ra_code)
        if not period_id:
            missing_period.append(f"{assign.ra_code} (no period for cycle {args.cycle})")
            continue
        pair = (mod["id"], period_id)
        if pair in existing_pairs:
            continue
        to_insert.append(
            {
                "module_id": mod["id"],
                "period_id": period_id,
                "status": "pending",
                "course_code": assign.course_code,
                "group_name": assign.group_name,
                "ra_code": assign.ra_code,
            }
        )

    extra_in_db: list[dict] = []
    excel_pairs: set[tuple[int, int]] = set()
    for assign in assignments:
        mod = module_by_key.get((assign.course_code, assign.group_name))
        period_id = period_by_ra.get(assign.ra_code)
        if mod and period_id:
            excel_pairs.add((mod["id"], period_id))

    for row in existing:
        pair = (row["module_id"], row["period_id"])
        if pair not in excel_pairs:
            extra_in_db.append(row)

    print(f"Cycle: {cycle['code']} (id={cycle['id']})")
    print(f"Modules in DB: {len(modules)}")
    print(f"Existing evaluations: {len(existing)}")
    print(f"Already present (Excel): {len(assignments) - len(to_insert) - len(missing_module) - len(missing_period)}")
    print(f"To insert: {len(to_insert)}")
    print(f"Insert by RA: {dict(Counter(r['ra_code'] for r in to_insert))}")
    print(f"Excel rows without DB module: {len(missing_module)}")
    print(f"Excel RAs without period: {len(missing_period)}")
    print(f"DB evaluations not in Excel: {len(extra_in_db)}")

    if missing_module[:10]:
        print("Sample missing modules:")
        for item in missing_module[:10]:
            print(f"  - {item}")
    if missing_period:
        print("Missing periods:")
        for item in sorted(set(missing_period)):
            print(f"  - {item}")

    if args.export_json:
        payload = {
            "cycle_code": cycle["code"],
            "source_excel": str(args.excel),
            "to_insert": to_insert,
            "missing_module": missing_module,
            "extra_in_db_count": len(extra_in_db),
        }
        args.export_json.parent.mkdir(parents=True, exist_ok=True)
        args.export_json.write_text(json.dumps(payload, indent=2, ensure_ascii=False), encoding="utf-8")
        print(f"Exported JSON: {args.export_json}")

    if missing_module or missing_period:
        return 1

    if args.dry_run:
        print("\n[dry-run] Sample inserts:")
        for row in to_insert[:15]:
            print(
                f"  {row['course_code']} · {row['group_name']} · {row['ra_code']} "
                f"→ module_id={row['module_id']} period_id={row['period_id']}"
            )
        return 0

    inserted = 0
    for row in to_insert:
        sb.table("module_ra_evaluations").insert(
            {
                "module_id": row["module_id"],
                "period_id": row["period_id"],
                "status": row["status"],
            }
        ).execute()
        inserted += 1

    final_count = len(existing) + inserted
    print(f"Inserted module_ra_evaluations: {inserted}")
    print(f"Expected total after sync: {final_count} (Excel target: {len(assignments)})")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

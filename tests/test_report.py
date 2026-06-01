from datetime import date
from decimal import Decimal
from io import BytesIO
from zipfile import ZipFile

import openpyxl
import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import StaticPool

from src.api.deps import get_db
from src.api.main import app
from src.core.security import hash_password
from src.db.base import Base
from src.models.action_plan import ActionPlan
from src.models.assessment import Assessment
from src.models.leader_analysis import LeaderAnalysis
from src.models.module import Module, ModuleAssignment
from src.models.module_analysis import ModuleAnalysis
from src.models.period import Period
from src.models.program import Program, ProgramMembership, PropedeuticLine
from src.models.rubric import PILevel, PerfIndicator, Rubric
from src.models.security import SecurityEvent
from src.models.student import ModuleStudent, Student
from src.models.student_outcome import StudentOutcome
from src.models.user import User

LOGIN_URL = "/api/v1/auth/login"
PREVIEW_URL = "/api/v1/periods/{}/report/preview"
EXPORT_URL = "/api/v1/periods/{}/report/export"
LEADER_REPORT_URL = "/api/v1/periods/{}/leader-report"
LEADER_REPORT_PDF_URL = "/api/v1/periods/{}/leader-report/pdf"
LEADER_REPORT_DOCX_URL = "/api/v1/periods/{}/leader-report/docx"
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
DOCX_MIME = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"


@pytest_asyncio.fixture
async def report_client():
    engine = create_async_engine(
        "sqlite+aiosqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    factory = async_sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    async with factory() as db:
        admin = User(
            email="admin.report@iub.edu.co",
            full_name="Admin Report",
            role="admin",
            hashed_password=hash_password("Admin1234!"),
            is_active=True,
            auth_provider="local",
        )
        leader = User(
            email="leader.report@iub.edu.co",
            full_name="Leader Report",
            role="leader",
            hashed_password=hash_password("Leader1234!"),
            is_active=True,
            auth_provider="local",
        )
        other_leader = User(
            email="other.leader.report@iub.edu.co",
            full_name="Other Leader Report",
            role="leader",
            hashed_password=hash_password("Leader1234!"),
            is_active=True,
            auth_provider="local",
        )
        teacher = User(
            email="teacher.report@iub.edu.co",
            full_name="=CMD Teacher",
            role="teacher",
            hashed_password=hash_password("Teacher1234!"),
            is_active=True,
            auth_provider="local",
        )
        db.add_all([admin, leader, other_leader, teacher])
        await db.flush()

        line = PropedeuticLine(code="LP-RPT", name="Linea Reporte", is_active=True)
        db.add(line)
        await db.flush()

        program = Program(
            propedeutic_line_id=line.id,
            name="Ingenieria de Reportes",
            code="IR-RPT",
            cycle_level="tecnologia",
            faculty="FIET",
        )
        other_program = Program(
            propedeutic_line_id=line.id,
            name="Otro Programa Reporte",
            code="OP-RPT",
            cycle_level="tecnologia",
            faculty="FIET",
        )
        db.add_all([program, other_program])
        await db.flush()

        db.add(ProgramMembership(user_id=leader.id, program_id=program.id, role="leader"))
        db.add(ProgramMembership(user_id=other_leader.id, program_id=other_program.id, role="leader"))
        await db.flush()

        so = StudentOutcome(
            code="RA-RPT",
            description="Resultado de Aprendizaje Reporte",
            is_active=True,
            program_id=program.id,
        )
        db.add(so)
        await db.flush()

        period = Period(
            name="2024-2 Reporte",
            student_outcome_id=so.id,
            start_date=date(2024, 8, 1),
            end_date=date(2024, 12, 1),
            status="closed",
            created_by=admin.id,
        )
        db.add(period)
        await db.flush()

        rubric = Rubric(student_outcome_id=so.id, period_id=period.id)
        db.add(rubric)
        await db.flush()

        pi1 = PerfIndicator(
            rubric_id=rubric.id,
            code="PI-RPT-1",
            description="Primer indicador de reporte",
            pi_weight=Decimal("60.00"),
            is_active=True,
            position=1,
        )
        pi2 = PerfIndicator(
            rubric_id=rubric.id,
            code="PI-RPT-2",
            description="Segundo indicador de reporte",
            pi_weight=Decimal("40.00"),
            is_active=True,
            position=2,
        )
        db.add_all([pi1, pi2])
        await db.flush()

        for pi in [pi1, pi2]:
            db.add_all(
                [
                    PILevel(perf_indicator_id=pi.id, level_value=1, label="Poor", descriptor="Bajo"),
                    PILevel(
                        perf_indicator_id=pi.id,
                        level_value=2,
                        label="Inadequate",
                        descriptor="Insuficiente",
                    ),
                    PILevel(
                        perf_indicator_id=pi.id,
                        level_value=3,
                        label="Adequate",
                        descriptor="Adecuado",
                    ),
                    PILevel(
                        perf_indicator_id=pi.id,
                        level_value=4,
                        label="Exemplary",
                        descriptor="Excelente",
                    ),
                ]
            )

        period.rubric_id = rubric.id

        module = Module(
            period_id=period.id,
            course_code="RPT101",
            course_name="Reportes I",
            group_name="G1",
            status="completed",
        )
        empty_module = Module(
            period_id=period.id,
            course_code="RPT000",
            course_name="Sin estudiantes",
            group_name="G0",
            status="pending",
        )
        db.add_all([module, empty_module])
        await db.flush()

        students = [
            Student(internal_id="RPT-1", document_number="9001", full_name="=SUM(A1)"),
            Student(internal_id="RPT-2", document_number="9002", full_name="Estudiante Normal"),
        ]
        db.add_all(students)
        await db.flush()

        module_students = [
            ModuleStudent(module_id=module.id, student_id=student.id, status="active")
            for student in students
        ]
        db.add_all(module_students)
        await db.flush()

        db.add_all(
            [
                Assessment(module_student_id=module_students[0].id, perf_indicator_id=pi1.id, level=4),
                Assessment(module_student_id=module_students[1].id, perf_indicator_id=pi1.id, level=3),
                Assessment(module_student_id=module_students[0].id, perf_indicator_id=pi2.id, level=2),
                Assessment(module_student_id=module_students[1].id, perf_indicator_id=pi2.id, level=2),
                ModuleAssignment(module_id=module.id, user_id=teacher.id),
                ModuleAssignment(module_id=empty_module.id, user_id=teacher.id),
                ModuleAnalysis(
                    module_id=module.id,
                    perf_indicator_id=pi1.id,
                    analysis_text="Analisis docente PI 1",
                ),
                ModuleAnalysis(
                    module_id=module.id,
                    perf_indicator_id=pi2.id,
                    analysis_text="Analisis docente PI 2",
                ),
                LeaderAnalysis(
                    period_id=period.id,
                    perf_indicator_id=pi1.id,
                    analysis_text="Sintesis lider PI 1",
                    updated_by=leader.id,
                ),
                LeaderAnalysis(
                    period_id=period.id,
                    perf_indicator_id=pi2.id,
                    analysis_text="Sintesis lider PI 2",
                    updated_by=leader.id,
                ),
                ActionPlan(
                    period_id=period.id,
                    perf_indicator_id=pi1.id,
                    action_type="improvement",
                    description="Fortalecer evidencias PI 1",
                    responsible="Comite RA",
                    estimated_date="2025-01",
                    updated_by=leader.id,
                ),
                ActionPlan(
                    period_id=period.id,
                    perf_indicator_id=pi2.id,
                    action_type="corrective",
                    description="Reforzar resultado PI 2",
                    responsible="Comite RA",
                    estimated_date="2025-02",
                    updated_by=leader.id,
                ),
            ]
        )
        await db.commit()

        ids = {"period_id": period.id, "pi1_id": pi1.id, "pi2_id": pi2.id}

    async def _override_get_db():
        async with factory() as session:
            try:
                yield session
            except Exception:
                await session.rollback()
                raise

    app.dependency_overrides[get_db] = _override_get_db
    transport = ASGITransport(app=app)
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client, ids, factory

    app.dependency_overrides.clear()
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)
    await engine.dispose()


async def _login(client: AsyncClient, email: str, password: str) -> None:
    resp = await client.post(LOGIN_URL, json={"email": email, "password": password})
    assert resp.status_code == 200


@pytest.mark.asyncio
async def test_leader_previews_report_with_four_sections(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(PREVIEW_URL.format(ids["period_id"]))

    assert resp.status_code == 200
    data = resp.json()
    assert data["period"]["name"] == "2024-2 Reporte"
    assert data["student_outcome"]["code"] == "RA-RPT"
    assert [module["course_code"] for module in data["modules_summary"]] == ["RPT101"]
    assert set(data["distribution_by_pi"]) == {"PI-RPT-1", "PI-RPT-2"}
    assert data["leader_analysis"]["PI-RPT-1"] == "Sintesis lider PI 1"
    assert data["action_plans"][0]["description"] == "Fortalecer evidencias PI 1"


@pytest.mark.asyncio
async def test_teacher_cannot_preview_report(report_client):
    client, ids, _factory = report_client
    await _login(client, "teacher.report@iub.edu.co", "Teacher1234!")

    resp = await client.get(PREVIEW_URL.format(ids["period_id"]))

    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_export_requires_complete_leader_analysis_and_action_plans(report_client):
    client, ids, factory = report_client
    async with factory() as db:
        await db.execute(
            delete(ActionPlan).where(ActionPlan.perf_indicator_id == ids["pi2_id"])
        )
        await db.commit()

    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(EXPORT_URL.format(ids["period_id"]), params={"format": "pdf"})

    assert resp.status_code == 409
    detail = resp.json()["detail"]
    assert detail["missing_action_plans"] == ["PI-RPT-2"]
    assert detail["missing_leader_analysis"] == []


@pytest.mark.asyncio
async def test_leader_exports_pdf_and_audit_event(report_client):
    client, ids, factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(EXPORT_URL.format(ids["period_id"]), params={"format": "pdf"})

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")

    async with factory() as db:
        result = await db.execute(
            select(SecurityEvent).where(SecurityEvent.event == "report_exported")
        )
        event = result.scalar_one()
        assert event.detail["period_id"] == ids["period_id"]
        assert event.detail["format"] == "pdf"


@pytest.mark.asyncio
async def test_xlsx_export_sanitizes_user_text_cells(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(EXPORT_URL.format(ids["period_id"]), params={"format": "xlsx"})

    assert resp.status_code == 200
    assert resp.headers["content-type"] == XLSX_MIME
    workbook = openpyxl.load_workbook(BytesIO(resp.content))
    values = [
        cell
        for sheet in workbook.worksheets
        for row in sheet.iter_rows(values_only=True)
        for cell in row
        if isinstance(cell, str)
    ]
    assert "'=SUM(A1)" in values
    assert "'=CMD Teacher" in values


@pytest.mark.asyncio
async def test_xlsx_export_includes_excel_parity_distribution_details(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(EXPORT_URL.format(ids["period_id"]), params={"format": "xlsx"})

    assert resp.status_code == 200
    workbook = openpyxl.load_workbook(BytesIO(resp.content))
    distribution = workbook["Distribucion"]
    rows = list(distribution.iter_rows(values_only=True))

    assert rows[0] == (
        "PI",
        "Descripcion PI",
        "Nivel",
        "Descriptor",
        "Modulo",
        "Porcentaje",
        "Conteo",
    )
    assert ("PI-RPT-1", "Primer indicador de reporte", "Exemplary", "Excelente", "RPT101", 50.0, 1) in rows
    assert (
        "PI-RPT-1",
        "Primer indicador de reporte",
        "Adequate",
        "Adecuado",
        "TOTAL CONSOLIDADO",
        50.0,
        1,
    ) in rows


@pytest.mark.asyncio
async def test_leader_report_preview_returns_metrics_and_empty_conclusions(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.get(LEADER_REPORT_URL.format(ids["period_id"]))

    assert resp.status_code == 200
    data = resp.json()
    assert data["period"]["name"] == "2024-2 Reporte"
    assert [item["pi_code"] for item in data["items"]] == ["PI-RPT-1", "PI-RPT-2"]
    assert data["items"][0]["leader_analysis"] == "Sintesis lider PI 1"
    assert data["items"][0]["conclusion_text"] == ""
    assert data["items"][0]["distribution"]["counts"]["Exemplary"] == 1


@pytest.mark.asyncio
async def test_leader_report_put_sanitizes_conclusions(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")

    resp = await client.put(
        LEADER_REPORT_URL.format(ids["period_id"]),
        json={
            "conclusions": [
                {
                    "perf_indicator_id": ids["pi1_id"],
                    "conclusion_text": "<script>alert(1)</script>=CMD(\"calc\")",
                }
            ]
        },
    )

    assert resp.status_code == 200
    data = resp.json()
    conclusion = data["items"][0]["conclusion_text"]
    assert "<script>" not in conclusion
    assert conclusion.endswith('=CMD("calc")')


@pytest.mark.asyncio
async def test_leader_report_exports_pdf_and_audits(report_client):
    client, ids, factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")
    await client.put(
        LEADER_REPORT_URL.format(ids["period_id"]),
        json={
            "conclusions": [
                {"perf_indicator_id": ids["pi1_id"], "conclusion_text": "Conclusion PI 1"},
                {"perf_indicator_id": ids["pi2_id"], "conclusion_text": "Conclusion PI 2"},
            ]
        },
    )

    resp = await client.get(LEADER_REPORT_PDF_URL.format(ids["period_id"]))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == "application/pdf"
    assert resp.content.startswith(b"%PDF")
    async with factory() as db:
        result = await db.execute(
            select(SecurityEvent).where(SecurityEvent.event == "leader_report_generated")
        )
        event = result.scalar_one()
        assert event.detail["period_id"] == ids["period_id"]
        assert event.detail["format"] == "pdf"


@pytest.mark.asyncio
async def test_leader_report_docx_sanitizes_formula_like_text(report_client):
    client, ids, _factory = report_client
    await _login(client, "leader.report@iub.edu.co", "Leader1234!")
    await client.put(
        LEADER_REPORT_URL.format(ids["period_id"]),
        json={
            "conclusions": [
                {"perf_indicator_id": ids["pi1_id"], "conclusion_text": '=CMD("calc")'},
                {"perf_indicator_id": ids["pi2_id"], "conclusion_text": "Conclusion segura"},
            ]
        },
    )

    resp = await client.get(LEADER_REPORT_DOCX_URL.format(ids["period_id"]))

    assert resp.status_code == 200
    assert resp.headers["content-type"] == DOCX_MIME
    with ZipFile(BytesIO(resp.content)) as docx:
        document_xml = docx.read("word/document.xml").decode("utf-8")
    assert "'=CMD(&quot;calc&quot;)" in document_xml


@pytest.mark.asyncio
async def test_teacher_cannot_use_leader_report(report_client):
    client, ids, _factory = report_client
    await _login(client, "teacher.report@iub.edu.co", "Teacher1234!")

    resp = await client.get(LEADER_REPORT_URL.format(ids["period_id"]))

    assert resp.status_code == 403
